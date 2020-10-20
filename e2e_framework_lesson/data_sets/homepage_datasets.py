import openpyxl
import os
from os.path import join, dirname
workbook_path = join(dirname(__file__), 'homepage_data.xlsx')


class HomePageData:

    test_homepage_data = [{
        "firstname": "Rahul",
        "lastname": "shetty",
        "gender": "Male"},
        {"firstname": "Anshika",
         "lastname": "shetty",
         "gender": "Female"}]

    @staticmethod
    def get_test_data(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook(workbook_path)
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]
